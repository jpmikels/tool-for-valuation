from flask import Flask, request, jsonify, send_file, render_template
from werkzeug.utils import secure_filename
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os
import re
from datetime import datetime
import tempfile

try:
    from google.cloud import documentai
    from google.api_core import exceptions as gcp_exceptions
    DOCUMENT_AI_AVAILABLE = True
except ImportError:
    DOCUMENT_AI_AVAILABLE = False

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['ALLOWED_EXTENSIONS'] = {'pdf'}

# GCP Configuration
USE_DOCUMENT_AI = os.environ.get('USE_DOCUMENT_AI', 'false').lower() == 'true'
PROJECT_ID = os.environ.get('GCP_PROJECT_ID', '')
PROCESSOR_ID = os.environ.get('DOCUMENT_AI_PROCESSOR_ID', '')
LOCATION = os.environ.get('GCP_LOCATION', 'us')

# Initialize Document AI client if available
document_ai_client = None
if DOCUMENT_AI_AVAILABLE and USE_DOCUMENT_AI:
    try:
        document_ai_client = documentai.DocumentProcessorServiceClient()
        print(f"✅ Document AI initialized for project: {PROJECT_ID}")
    except Exception as e:
        print(f"⚠️  Document AI initialization failed: {str(e)}")
        document_ai_client = None

# Ensure upload folder exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def extract_with_document_ai(pdf_path):
    """Extract financial data using GCP Document AI"""
    if not document_ai_client or not PROJECT_ID or not PROCESSOR_ID:
        return None
    
    try:
        # Read the PDF file
        with open(pdf_path, 'rb') as f:
            file_content = f.read()
        
        # Configure the process request
        name = f"projects/{PROJECT_ID}/locations/{LOCATION}/processors/{PROCESSOR_ID}"
        
        # Prepare the request
        request = documentai.types.ProcessRequest(
            name=name,
            raw_document=documentai.RawDocument(
                content=file_content,
                mime_type="application/pdf"
            )
        )
        
        # Process the document
        result = document_ai_client.process_document(request=request)
        document = result.document
        
        # Extract tables from the document
        tables_data = []
        for page in document.pages:
            for table in page.tables:
                table_rows = []
                for row in table.header_rows:
                    row_data = []
                    for cell in row.cells:
                        cell_text = []
                        for layout in cell.layout.text_anchor.text_segments:
                            start_index = layout.start_index
                            end_index = layout.end_index
                            cell_text.append(document.text[start_index:end_index])
                        row_data.append(' '.join(cell_text))
                    table_rows.append(row_data)
                
                for row in table.body_rows:
                    row_data = []
                    for cell in row.cells:
                        cell_text = []
                        for layout in cell.layout.text_anchor.text_segments:
                            start_index = layout.start_index
                            end_index = layout.end_index
                            cell_text.append(document.text[start_index:end_index])
                        row_data.append(' '.join(cell_text))
                    table_rows.append(row_data)
                
                if table_rows:
                    tables_data.append(table_rows)
        
        return tables_data
    except Exception as e:
        print(f"Document AI processing error: {str(e)}")
        return None

def extract_financial_data_pdfplumber(pdf_path):
    """Extract financial data from PDF using pdfplumber (fallback)"""
    data = {
        'pl': [],
        'income': [],
        'cashflow': []
    }
    
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue
                
            # Try to extract tables
            tables = page.extract_tables()
            for table in tables:
                if table:
                    cleaned_table = []
                    for row in table:
                        if row:
                            cleaned_row = [str(cell).strip() if cell else '' for cell in row]
                            if any(cleaned_row):
                                cleaned_table.append(cleaned_row)
                    
                    if cleaned_table:
                        header = cleaned_table[0] if cleaned_table else []
                        header_text = ' '.join(header).lower()
                        
                        if any(keyword in header_text for keyword in ['profit', 'loss', 'revenue', 'income', 'expense']):
                            if 'cash' in header_text or 'flow' in header_text:
                                data['cashflow'].extend(cleaned_table)
                            elif 'profit' in header_text or 'loss' in header_text or 'revenue' in header_text:
                                data['pl'].extend(cleaned_table)
                            else:
                                data['income'].extend(cleaned_table)
                        else:
                            all_text = text.lower()
                            if 'net income' in all_text or 'revenue' in all_text:
                                data['income'].extend(cleaned_table)
                            elif 'operating activities' in all_text or 'financing' in all_text:
                                data['cashflow'].extend(cleaned_table)
    
    return data

def classify_and_extract_data(tables_data):
    """Classify tables and organize financial data"""
    organized = {
        'income': [],
        'pl': [],
        'cashflow': []
    }
    
    for table in tables_data:
        if not table:
            continue
            
        header = table[0] if table else []
        header_text = ' '.join([str(cell).lower() for cell in header])
        
        # Determine statement type
        if any(keyword in header_text for keyword in ['cash', 'operating', 'financing', 'investing']):
            organized['cashflow'].extend(table)
        elif any(keyword in header_text for keyword in ['revenue', 'income', 'expense', 'profit', 'loss', 'net']):
            organized['income'].extend(table)
        else:
            organized['pl'].extend(table)
    
    return organized

def extract_financial_data(pdf_path):
    """Main extraction function - tries Document AI first, falls back to pdfplumber"""
    # Try Document AI first
    if USE_DOCUMENT_AI and document_ai_client:
        ai_tables = extract_with_document_ai(pdf_path)
        if ai_tables and ai_tables:
            print("📊 Using GCP Document AI for extraction")
            return classify_and_extract_data(ai_tables)
    
    # Fallback to pdfplumber
    print("📄 Using pdfplumber for extraction")
    return extract_financial_data_pdfplumber(pdf_path)

def create_excel_workbook(data):
    """Create Excel workbook with financial data"""
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center")
    
    # Create sheets
    sheets_data = [
        ('Income Statement', data.get('income', []) + data.get('pl', [])),
        ('Cash Flow Statement', data.get('cashflow', []))
    ]
    
    for sheet_name, sheet_data in sheets_data:
        if sheet_data:
            ws = wb.create_sheet(title=sheet_name)
            ws.append([''])  # Empty row
            
            # Add data
            for row in sheet_data:
                ws.append(row)
            
            # Apply header styling to first row with data
            if len(sheet_data) > 0:
                for cell in ws[2]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    return wb

@app.route('/')
def index():
    ai_status = {
        'document_ai_enabled': USE_DOCUMENT_AI and document_ai_client is not None,
        'library_available': DOCUMENT_AI_AVAILABLE
    }
    return render_template('index.html', ai_status=ai_status)

@app.route('/upload', methods=['POST'])
def upload_files():
    """Handle PDF file uploads and process them"""
    if 'files[]' not in request.files:
        return jsonify({'error': 'No files provided'}), 400
    
    files = request.files.getlist('files[]')
    if not files or files[0].filename == '':
        return jsonify({'error': 'No files selected'}), 400
    
    all_data = {
        'pl': [],
        'income': [],
        'cashflow': []
    }
    
    processed_files = []
    
    for file in files:
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], f'{timestamp}_{filename}')
            file.save(filepath)
            
            try:
                extracted_data = extract_financial_data(filepath)
                
                # Merge data
                all_data['pl'].extend(extracted_data['pl'])
                all_data['income'].extend(extracted_data['income'])
                all_data['cashflow'].extend(extracted_data['cashflow'])
                
                processed_files.append(filename)
                
            except Exception as e:
                print(f"Error processing {filename}: {str(e)}")
            finally:
                # Clean up
                if os.path.exists(filepath):
                    os.remove(filepath)
    
    if not any(all_data.values()):
        return jsonify({'error': 'No financial data extracted from PDFs'}), 400
    
    # Create Excel workbook
    try:
        wb = create_excel_workbook(all_data)
        
        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        # Return file
        return send_file(
            temp_file.name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='financial_statements_merged.xlsx'
        )
    except Exception as e:
        return jsonify({'error': f'Error creating Excel file: {str(e)}'}), 500

@app.route('/health')
def health():
    status = {
        'status': 'healthy',
        'ai_enabled': USE_DOCUMENT_AI,
        'document_ai_available': document_ai_client is not None
    }
    return jsonify(status), 200

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=int(os.environ.get('PORT', 8080)))
